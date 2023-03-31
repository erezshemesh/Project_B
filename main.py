"""
    This file is the executable for running PPO. It is based on this medium article:
    https://medium.com/@eyyu/coding-ppo-from-scratch-with-pytorch-part-1-4-613dfc1b14c8
"""
import math

import gym
import sys
import torch
from train_system import *
from arguments import get_args
from ppo import PPO
from network import FeedForwardNN
from eval_policy import eval_policy
import openpyxl
import pandas as pd
from openpyxl import load_workbook
import os
from torch.utils.tensorboard import SummaryWriter

# consts:
DAY_BEGIN_SEC = 21600
DAY_END_SEC = 86400

# Table generation:

T = np.array([[41002, 43133, 45466, 47962],
              [76220, 79164, 82408, 85859],
              [82260, 83589, 84864, 86039],
              [82917, 84022, 85122, 86219],
              [83125, 84216, 85307, 86399]])

L = np.array([[1940, 3209, 4096, 4770],
              [3521, 5716, 7123, 8063],
              [603, 804, 728, 455],
              [65, 82, 75, 63],
              [20, 31, 37, 40]])

P = np.array([[1940, 2045, 2170, 2312],
              [3521, 3603, 3694, 3789],
              [603, 442, 245, 18],
              [65, 43, 25, 18],
              [20, 19, 18, 18]])


# This function creates a dictionary from the first two columns of specified workbook sheet
def dic_creator(sheet_name, workbook, dic):
    worksheet = workbook[sheet_name]
    for row in worksheet.iter_rows(values_only=True, min_row=2):
        key = row[0]
        value = row[1]
        dic[key] = value


def update_history_counter(max_history, reset_en=False):
    # gets current value of history_counter:
    current_value = counter_worksheet['B2'].value
    if current_value == max_history or reset_en:
        new_value = 1
    else:
        new_value = current_value + 1
    # write the new value back to the cell
    counter_worksheet['B2'] = new_value
    config_workbook.save('config_file.xlsx')


def save_table(table, table_type, folder_name):
    # specify the filename to write to with a variable holding an integer value
    filename = table_type + "_" + str(history_counter) + ".txt"
    # create the full path to the file
    path = os.path.join(os.getcwd(), folder_name, filename)
    # check if the file exists
    if os.path.isfile(path):
        # if the file exists, overwrite it with the new tensor data
        np.savetxt(path, table)
        print((table_type + "_" + str(history_counter)), " overwritten successfully!")
    else:
        # if the file does not exist, create a new one and write the table to it
        # create the directory if it does not exist
        os.makedirs(os.path.join(os.getcwd(), folder_name), exist_ok=True)
        np.savetxt(path, table)
        print((table_type + "_" + str(history_counter)), " created and written successfully!")

def save_generated_tables(T,L,P):
    save_table(T, "T", "T_history")
    save_table(L, "L", "L_history")
    save_table(P, "P", "P_history")


def table_loader(en, history_index=math.nan):  # TODO: instead of en, use command line
    if en:
        # Generate new tables:
        constraints = g.make_all_constraints
        g.sol = minimize(g.objective_max_board, g.V, method='SLSQP', constraints=constraints, callback=g.callback,
                         options={'maxiter': max_iterations})
        new_T, new_L, new_P = g.extract(g.sol.x)
        # Save tables to .txt files in directories:
        save_generated_tables(new_T, new_L, new_P)
        # History counter proceeding:
        update_history_counter(5) #TODO: make it paramater in config_file
        # Return new tables:
        return new_T, new_L, new_P
    else:
        #Load tables from history:
        filename_T = "T" + "_" + str(history_index) + ".txt"
        filename_L = "L" + "_" + str(history_index) + ".txt"
        filename_P = "P" + "_" + str(history_index) + ".txt"
        path_T = os.path.join(os.getcwd(), "T_history", filename_T)
        print (path_T)
        path_L = os.path.join(os.getcwd(), "L_history", filename_L)
        path_P = os.path.join(os.getcwd(), "P_history", filename_P)
        # Check file paths:
        if os.path.isfile(path_T) and os.path.isfile(path_L) and os.path.isfile(path_P):
            new_T = np.loadtxt(path_T)
            new_L = np.loadtxt(path_L)
            new_P = np.loadtxt(path_P)
        else:
            raise ValueError("One or more paths are not found! either folder or txt are missing!"
                             " make sure you stored tables before!")
        # Check shapes of tensors - TODO: if shape is not trains X stations - raise error
    return new_T, new_L, new_P


# handle config_file:
config_workbook = openpyxl.load_workbook('config_file.xlsx')
# creates a dictionary of hyperparameters:
hyper_param_dic = {}
dic_creator('hyperparameters', config_workbook, hyper_param_dic)
# creates a dictionary of generator parameters:
gen_param_dic = {}
dic_creator('generator_parameters', config_workbook, gen_param_dic)
# loads / generates tables:
counter_worksheet = config_workbook['counters']
history_counter = counter_worksheet['B2'].value
# if False:
#     update_history_counter(5)
# saves generated tables in history xlsx:
# closes workbook:
config_workbook.close()
# TODO: handle table saving according to flags: (Use chatGPT  \ Stackoverflow for help)


step_size = hyper_param_dic['step_size']
steps_per_day = int((DAY_END_SEC - DAY_BEGIN_SEC) / step_size)
days_in_episode = hyper_param_dic['days_in_episode']  # default value = 1
steps_per_episode = steps_per_day * days_in_episode
iteration_num = hyper_param_dic['iteration_number']  # default value = 1
# Batch can contain more than 1 episode as far as I understand
episodes_in_batch = hyper_param_dic['episodes_in_batch']

# Generator:

g = Generator(
    trains=gen_param_dic['trains'],
    stations=gen_param_dic['stations'],
    t_alight_per_person=gen_param_dic['t_alight_per_person'],
    t_board_per_person=gen_param_dic['t_board_per_person'],
    platform_arrivals_per_t=gen_param_dic['platform_arrivals_per_t'],
    alight_fraction=gen_param_dic['alight_fraction'],
    number_of_carts=gen_param_dic['number_of_carts'],
    km_between_stations=gen_param_dic['km_between_stations'],
    speed_kmh=gen_param_dic['speed_kmh'],
    stop_t=gen_param_dic['stop_t'],
    tmin=gen_param_dic['tmin'],
    train_capacity=gen_param_dic['train_capacity'],
    platform_capacity=gen_param_dic['platform_capacity'],
    var=gen_param_dic['var']
)


def train(env, hyperparameters, actor_model, critic_model):
    """
        Trains the model.

        Parameters:
            env - the environment to train on
            hyperparameters - a dict of hyperparameters to use, defined in main
            actor_model - the actor model to load in if we want to continue training
            critic_model - the critic model to load in if we want to continue training

        Return:
            None
    """
    print(f"Training", flush=True)

    # Create a model for PPO.
    model = PPO(policy_class=FeedForwardNN, env=env, **hyperparameters)

    # Tries to load in an existing actor/critic model to continue training on
    if actor_model != '' and critic_model != '':
        print(f"Loading in {actor_model} and {critic_model}...", flush=True)
        model.actor.load_state_dict(torch.load(actor_model))
        model.critic.load_state_dict(torch.load(critic_model))
        print(f"Successfully loaded.", flush=True)
    elif actor_model != '' or critic_model != '':  # Don't train from scratch if user accidentally forgets actor/critic model
        print(
            f"Error: Either specify both actor/critic models or none at all. We don't want to accidentally override anything!")
        sys.exit(0)
    else:
        print(f"Training from scratch.", flush=True)

    # Train the PPO model with a specified total times
    # NOTE: You can change the total timesteps here, I put a big number just because
    # you can kill the process whenever you feel like PPO is converging
    total_timesteps = steps_per_day * days_in_episode * episodes_in_batch * iteration_num
    model.learn(total_timesteps)

def test(env, actor_model):
    """
        Tests the model.

        Parameters:
            env - the environment to test the policy on
            actor_model - the actor model to load in

        Return:
            None
    """
    print(f"Testing {actor_model}", flush=True)

    # If the actor model is not specified, then exit
    if actor_model == '':
        print(f"Didn't specify model file. Exiting.", flush=True)
        sys.exit(0)

    # Extract out dimensions of observation and action spaces
    obs_dim = env.observation_space.shape[0]
    act_dim = env.action_space.shape[0]

    # Build our policy the same way we build our actor model in PPO
    policy = FeedForwardNN(obs_dim, act_dim)

    # Load in the actor model saved by the PPO algorithm
    policy.load_state_dict(torch.load(actor_model))

    # Evaluate our policy with a separate module, eval_policy, to demonstrate
    # that once we are done training the model/policy with ppo.py, we no longer need
    # ppo.py since it only contains the training algorithm. The model/policy itself exists
    # independently as a binary file that can be loaded in with torch.
    eval_policy(policy=policy, env=env, render=False)

def main(args):
    """
        The main function to run.

        Parameters:
            args - the arguments parsed from command line

        Return:
            None
    """

    # NOTE: Here's where you can set hyperparameters for PPO. I don't include them as part of
    # ArgumentParser because it's too annoying to type them every time at command line. Instead, you can change them here.
    # To see a list of hyperparameters, look in ppo.py at function _init_hyperparameters
    hyperparameters = {

        'timesteps_per_batch': steps_per_episode * episodes_in_batch,
        'max_timesteps_per_episode': steps_per_episode,
        'gamma': hyper_param_dic['gamma'],
        'n_updates_per_iteration': hyper_param_dic['n_updates_per_iteration'],
        'lr': hyper_param_dic['lr'],
        'clip': hyper_param_dic['clip'],
        'render': hyper_param_dic['render'],
        'render_every_i': hyper_param_dic['render_every_i']
    }

    # Creates the environment we'll be running. If you want to replace with your own
    # custom environment, note that it must inherit Gym and have both continuous
    # observation and action sp
    # aces.

    # env = gym.make('Pendulum-v0')
    env = GymTrainSystem(T, L, P, g, step_size)

    # Train or test, depending on the mode specified
    if args.mode == 'train':
        train(env=env, hyperparameters=hyperparameters, actor_model=args.actor_model, critic_model=args.critic_model)
    else:
        test(env=env, actor_model=args.actor_model)

if __name__ == '__main__':
    # If you manually delete history folders or some files inside, it is important to reset history_counter!
    #update_history_counter(5,True)
    T,L,P = table_loader(gen_param_dic['generate_new'], gen_param_dic['load_index'])
    args = get_args()  # Parse arguments from command line
    main(args)
    print(T,"\n")
    print(L,"\n")
    print(P,"\n")
